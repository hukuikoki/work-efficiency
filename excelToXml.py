import openpyxl
import pathlib

# エクセルファイルの取り込み
wb = openpyxl.load_workbook("./ブック名.xlsx")
yoyakuSheet = wb["sheet1"]
sokuhaiSheet = wb["sheet2"]
# 書き出すファイルのｎパス
inventoryFile = pathlib.Path('ファイルパス')

# EXCELのテーブルの値をハッシュ化して配列に追加
def tableToList(ws):
    # １行目（列名のセル）
    header_cells = ws[1]
    # ２行目以降（データ）
    # 在庫リスト
    inventory_list = []
    for row in ws.iter_rows(min_row=2):
        row_dic = {}
        # セルの値を「key-value」で登録
        for k, v in zip(header_cells, row):
            row_dic[k.value] = v.value
        inventory_list.append(row_dic)

    return inventory_list

# xml宣言とルート開始タグを書き込み
def writeDeclaration():
    with open(inventoryFile, 'w') as f:
        print('<?xml version="1.0" encoding="UTF-8"?>\n<inventory xmlns="http://www.demandware.com/xml/impex/inventory/2007-05-31">', file=f)
# 主要データの書き込み
def writeMain(inventoryList, isYoyaku):
    with open(inventoryFile, 'a') as f:
        print('    <xxxxxx>', file=f)
        print('        <xxxxxx>', file=f) if isYoyaku == True  else print('        <xxxxxx>', file=f)
        print('            <xxxxxx>huga</xxxxxx>', file=f)
        print('            <xxxxxx>hoge</xxxxxx>', file=f) if isYoyaku == True else print('        <xxxxxx>hoge2</xxxxxx>', file=f)
        print('            <xxxxxx>huga</xxxxxx>', file=f)
        print('            <xxxxxx>hoge</xxxxxx>', file=f)
        print('        </xxxxxx>\n', file=f)
        print('        <xxxxxx>', file=f)

        for inventory in inventoryList:
            print('            <xxxxxx="' + str(inventory['productId'])  + '">', file=f)
            print('                <xxxxxx>' + str(inventory['allocation'])  + '</xxxxxx>', file=f)
            print('            </xxxxxx=>', file=f)

        print('        </xxxxxx>', file=f)
        print('    </xxxxxx>', file=f)
# ルート終了タグを書き込みを書き込み
def writeEnd():
    with open(inventoryFile, 'a') as f:
        print('</xxxxxxx>', file=f)

# 実行
writeDeclaration()
yoyakuInventoryList = tableToList(yoyakuSheet)
writeMain(yoyakuInventoryList, True)
sokuhaiInventoryList = tableToList(sokuhaiSheet)
writeMain(sokuhaiInventoryList, False)
writeEnd()