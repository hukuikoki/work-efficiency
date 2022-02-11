import glob
import os
import openpyxl
 
#①対象ファイルのパス
path = '../excel'
  
#②対象ファイル種別
fileType = '*.xlsx'
  
#③置換対象としたいシート名
sheetName = ['表紙']
  
#④置換対象項目名
tgtItem = ['置き換え対象データ']
  
#⑤置換後データ
changDate = '置き換え後データ'
#「①対象ファイルのパス」配下にあるExcelファイルのパスを出力
print("■検索対象ファイル")
print(glob.glob(os.path.join(path,fileType )))
  
#「①対象ファイルのパス」配下にある「xlsx」ファイル数分ループ
for book in glob.glob(os.path.join(path, fileType)):
   print("■対象ファイル")
   print(book)
   bookFlg=0
   #ブックの取得
   #openpyxl.load_workbook('Excelファイルのパス')
   actBook = openpyxl.load_workbook(book)
  
   #シート数分ループ
   for actSheetName in actBook.sheetnames:
  
      print("■対象シート")
      print(actSheetName)
      count = 0

      #シート名の判定(「③置換対象としたいシート名」との比較)
      if actSheetName in sheetName:
      
         #該当シートの最大行を取得
         maxRow = actBook[actSheetName].max_row
           
         #アクティブシートを取得
         #ブック変数[シート名]
         actSheet = actBook[actSheetName]
         #置換対象項目行のループ
         #for 行変数 in シート変数.iter_rows(min_row = 2(開始行)　,max_row=2(終了行))
         for row in actSheet.iter_rows(min_row=10,max_row=10):
     
            #for セル変数 in 行変数
            for cellRow in row:
              
               #置換対象項目の判定(「④置換対象項目名」との比較)
               if cellRow.value in tgtItem:
               #セル変数.value=置換文字
                  cellRow.value=changDate
                  count+=1
                  bookFlg=1
                           
      print(str(count) + "件置換しました。")
        
   #ブックを保存
   if bookFlg == 1:
      #ブック変数.save(Excelファイルのパス)
      actBook.save(book)
   else:
      actBook.close
